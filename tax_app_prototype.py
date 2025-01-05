import streamlit as st
import pandas as pd
from openpyxl import load_workbook

# Load Excel data
def load_excel(file_path):
    excel_data = pd.ExcelFile(file_path)
    return excel_data

# Display sheet data
def display_sheet(excel_data, sheet_name):
    data = pd.read_excel(excel_data, sheet_name=sheet_name)
    st.write(f"### Data from Sheet: {sheet_name}")
    st.dataframe(data)
    return data

# Extract formulas from Excel sheet
def extract_formulas(file_path, sheet_name):
    wb = load_workbook(file_path, data_only=False)
    sheet = wb[sheet_name]

    formulas = {}
    for row in sheet.iter_rows():
        for cell in row:
            if cell.formula:  # Check if the cell contains a formula
                formulas[cell.coordinate] = cell.formula

    return formulas

# Display formulas
def display_formulas(file_path, sheet_name):
    st.write(f"### Formulas in Sheet: {sheet_name}")
    formulas = extract_formulas(file_path, sheet_name)

    if formulas:
        for cell, formula in formulas.items():
            st.write(f"Cell {cell}: {formula}")
    else:
        st.write("No formulas found in this sheet.")

# Calculate Tax with Brackets

def calculate_tax_with_brackets(data):
    # Income Sources and Codes
    income_sources = {
        "Income/(loss) from property": "2000",
        "Rent received or receivable": "2001",
        "1/10th of amount not adjustable against rent": "2002",
        "Forfeited deposit under a contract for sale of property": "2003",
        "Recovery of Unpaid Irrecoverable Rent allowed as deduction": "2004",
        "Income from Business": "3000",
        "Gains / (Loss) from Capital Assets (including securities)": "4000",
        "Income / (Loss) from Other Sources": "5000",
        "Foreign Income": "6000",
        "Agriculture Income": "6100",
    }

    deductions = {
        "Zakat u/s 60": "9001",
        "Workers Welfare Fund u/s 60A": "9002",
        "Educational expenses u/s 60D": "9008",
        "Import u/s 148 @1%": "64010052",
        "Import u/s 148 @2%": "64010054",
        "Import u/s 148 @3%": "64010056",
        "Dividend u/s 150 @7.5%": "64030052",
        "Yield on Behbood Certificates": "64030071",
        "Profit on Debt u/s 7B": "64310056",
        "Capital Gains on Securities u/s 37A @ 15%": "64220156",
    }

    computations = {
        "Taxable Income": "9100",
        "Normal Income Tax": "9200",
        "Tax Credits": "9329",
    }

    # Extract Income Values
    total_income = 0
    for source, code in income_sources.items():
        if code in data.columns:
            total_income += data[code].sum()

    # Extract Deduction Values
    total_deductions = 0
    for deduction, code in deductions.items():
        if code in data.columns:
            total_deductions += data[code].sum()

    # Calculate Taxable Income
    taxable_income = total_income - total_deductions

    # Define Tax Brackets (Example)
    tax_brackets = [
        (0, 50000, 0.05),
        (50001, 100000, 0.10),
        (100001, 200000, 0.15),
        (200001, float('inf'), 0.20),
    ]

    tax = 0
    for bracket in tax_brackets:
        lower, upper, rate = bracket
        if taxable_income > lower:
            taxable_income_in_bracket = min(taxable_income, upper) - lower
            tax += taxable_income_in_bracket * rate

    st.write("### Tax Calculation")
    st.write(f"Total Income: {total_income}")
    st.write(f"Total Deductions: {total_deductions}")
    st.write(f"Taxable Income: {taxable_income}")
    st.write(f"Total Tax Payable: {tax}")

# Main Streamlit App
def main():
    st.title("Income Tax Assessment Tool")

    # File Upload
    uploaded_file = st.file_uploader("Upload your Excel file", type=["xlsx"])

    if uploaded_file is not None:
        excel_data = load_excel(uploaded_file)

        # Sheet Selection
        sheet_names = excel_data.sheet_names
        selected_sheet = st.selectbox("Select a sheet to view", sheet_names)

        if selected_sheet:
            data = display_sheet(excel_data, selected_sheet)

            # Display Formulas
            if st.button("Show Formulas"):
                display_formulas(uploaded_file, selected_sheet)

            # Tax Calculation
            if st.button("Calculate Tax"):
                calculate_tax_with_brackets(data)

    else:
        st.write("Please upload an Excel file to proceed.")

if __name__ == "__main__":
    main()
